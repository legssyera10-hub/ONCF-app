from __future__ import annotations

from datetime import date, datetime, time, timedelta, timezone
from io import BytesIO
import json
import re
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query, status
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from sqlalchemy import delete, or_, select
from sqlalchemy.orm import Session, joinedload
from starlette.responses import StreamingResponse

from app.api.deps import require_roles
from app.core.security import get_password_hash
from app.core.technicentres import TECHNICENTRE_CODES
from app.db.session import get_db
from app.models.alert import (
    Alert,
    AlertAttachment,
    AlertRevision,
    AlertStatusHistory,
    EstablishmentConfirmation,
    MailEvent,
    Notification,
    PermanentDecision,
)
from app.models.app_setting import AppSetting
from app.models.establishment import Establishment
from app.models.enums import UserRole
from app.models.online_trial import (
    OnlineTrialAttachment,
    OnlineTrialDecision,
    OnlineTrialRequest,
    OnlineTrialStatusHistory,
)
from app.models.station import Station
from app.models.user import User
from app.services.alert_form_config import get_alert_form_config, save_alert_form_config
from app.services.alerts import apply_alert_collection_derived_fields, compute_delay_minutes
from app.services.mailing import send_system_mail
from app.services.realtime import manager
from app.schemas.admin import (
    AdminAlertFormConfigRead,
    AdminAlertFormConfigUpdate,
    AdminAlertFormFieldConfig,
    AdminEstablishmentCreate,
    AdminEstablishmentCreateResponse,
    AdminMailRoutingSettingsRead,
    AdminMailRoutingTestPayload,
    AdminMailRoutingSettingsUpdate,
    AdminEstablishmentUpdate,
    AdminPasswordUpdate,
    AdminUserActivity,
    AdminUserCreate,
    AdminUserDetail,
    AdminUserRead,
    AdminUserUpdate,
    AdminStationCreate,
    AdminStationResponse,
    AdminStationUpdate,
)
from app.schemas.common import StationRead
from pydantic import BaseModel

router = APIRouter(prefix="/admin", tags=["admin"])


class AdminActionResponse(BaseModel):
    message: str


def _get_setting(db: Session, key: str) -> Optional[str]:
    setting = db.get(AppSetting, key)
    if not setting:
        return None
    value = (setting.value or "").strip()
    return value or None


def _set_setting(db: Session, key: str, value: Optional[str]) -> None:
    normalized = (value or "").strip()
    setting = db.get(AppSetting, key)
    if setting:
        setting.value = normalized
        return
    db.add(AppSetting(key=key, value=normalized))


def _describe_status_history(status_value: str, note: Optional[str]) -> str:
    label_map = {
        "EN_COURS_DE_TRAITEMENT": "En cours de traitement",
        "A_MODIFIER": "À modifier",
        "MODIFIEE": "Modifiée",
        "TRAITEE_PAR_PM": "Validée par le PPM",
        "ANNULEE": "Annulée",
        "RECEPTION_PARTIELLE": "Réception partielle",
        "RECEPTION_COMPLETE": "Réception confirmée",
    }
    label = label_map.get(status_value, f"Statut {status_value}")
    return label + (f" - {note}" if note else "")


def _status_business_label(status_value: str) -> str:
    return _describe_status_history(status_value, None)


def _slugify_establishment_code(value: str) -> str:
    normalized = (
        value.strip()
        .upper()
        .replace("'", " ")
        .replace("-", " ")
        .replace("+", " ")
    )
    parts = [part for part in normalized.split() if part]
    return "_".join(parts)[:50]


def _slugify_station_code(value: str) -> str:
    normalized = (
        value.strip()
        .upper()
        .replace("'", " ")
        .replace("-", " ")
        .replace("+", " ")
    )
    parts = [part for part in normalized.split() if part]
    return "_".join(parts)[:50]


def _validate_establishment_assignment(db: Session, role: UserRole, establishment_id: Optional[int]) -> None:
    if role == UserRole.ETABLISSEMENT:
        if not establishment_id:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Un compte etablissement doit etre lie a un etablissement",
            )
        if not db.get(Establishment, establishment_id):
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Etablissement introuvable")
        return
    if establishment_id is not None:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Seuls les comptes etablissement peuvent etre lies a un etablissement",
        )


def _user_activity_history(db: Session, user_id: int) -> list[AdminUserActivity]:
    user = db.get(User, user_id)
    if not user:
        return []

    if user.role == UserRole.SUIVI:
        tracked_alerts = _alerts_for_user(db, user)
        tracked_trials = _online_trials_for_user(db, user)
        timeline = [
            AdminUserActivity(
                timestamp=alert.updated_at or alert.created_at,
                action=alert.status.value,
                details=(
                    f"Acheminement {alert.material_ref} - {alert.station.name if alert.station else ''}"
                    + (
                        f" vers {alert.permanent_decision.destination_establishment.name}"
                        if alert.permanent_decision and alert.permanent_decision.destination_establishment
                        else ""
                    )
                ).strip(),
                alert_id=alert.id,
            )
            for alert in tracked_alerts
        ]
        timeline.extend(
            AdminUserActivity(
                timestamp=trial.updated_at or trial.created_at,
                action=trial.status.value,
                details=f"Essai en ligne {trial.material_ref} - {trial.station.name if trial.station else ''}".strip(),
                alert_id=None,
            )
            for trial in tracked_trials
        )
        timeline.sort(key=lambda item: item.timestamp, reverse=True)
        return timeline

    history: list[AdminUserActivity] = []

    alerts = list(db.execute(select(Alert).where(Alert.created_by_user_id == user_id)).scalars())
    for alert in alerts:
        history.append(
            AdminUserActivity(
                timestamp=alert.created_at,
                action="ALERTE_CREEE",
                details=f"Alerte {alert.material_ref} creee a {alert.station_id}",
                alert_id=alert.id,
            )
        )
    online_trials = list(db.execute(select(OnlineTrialRequest).where(OnlineTrialRequest.created_by_user_id == user_id)).scalars())
    for trial in online_trials:
        history.append(
            AdminUserActivity(
                timestamp=trial.created_at,
                action="ESSAI_CREE",
                details=f"Demande essai {trial.material_ref} creee a {trial.station_id}",
                alert_id=None,
            )
        )

    status_entries = list(
        db.execute(select(AlertStatusHistory).where(AlertStatusHistory.changed_by_user_id == user_id)).scalars()
    )
    for item in status_entries:
        history.append(
            AdminUserActivity(
                timestamp=item.changed_at,
                action="STATUT_MIS_A_JOUR",
                details=_describe_status_history(item.status.value, item.note),
                alert_id=item.alert_id,
            )
        )
    trial_status_entries = list(
        db.execute(select(OnlineTrialStatusHistory).where(OnlineTrialStatusHistory.changed_by_user_id == user_id)).scalars()
    )
    for item in trial_status_entries:
        history.append(
            AdminUserActivity(
                timestamp=item.changed_at,
                action="STATUT_ESSAI_MIS_A_JOUR",
                details=_describe_status_history(item.status.value, item.note),
                alert_id=None,
            )
        )

    decisions = list(db.execute(select(PermanentDecision).where(PermanentDecision.permanent_user_id == user_id)).scalars())
    for decision in decisions:
        history.append(
            AdminUserActivity(
                timestamp=decision.created_at,
                action="DECISION_PERMANENT",
                details=f"Decision {decision.decision.value} vers etablissement {decision.destination_establishment_id}",
                alert_id=decision.alert_id,
            )
        )
    trial_decisions = list(
        db.execute(select(OnlineTrialDecision).where(OnlineTrialDecision.permanent_user_id == user_id)).scalars()
    )
    for decision in trial_decisions:
        history.append(
            AdminUserActivity(
                timestamp=decision.updated_at or decision.created_at,
                action="DECISION_ESSAI_PERMANENT",
                details=f"Decision essai {decision.decision.value}",
                alert_id=None,
            )
        )

    confirmations = list(
        db.execute(
            select(EstablishmentConfirmation).where(EstablishmentConfirmation.establishment_user_id == user_id)
        ).scalars()
    )
    for confirmation in confirmations:
        history.append(
            AdminUserActivity(
                timestamp=confirmation.confirmed_at,
                action="RECEPTION_COMPLETE",
                details=confirmation.remarks or "Reception confirmee",
                alert_id=confirmation.alert_id,
            )
        )

    history.sort(key=lambda item: item.timestamp, reverse=True)
    return history


def _date_range(
    start_date: Optional[str],
    end_date: Optional[str],
) -> tuple[Optional[datetime], Optional[datetime], str]:
    if not start_date and not end_date:
        return None, None, "toutes_periodes"

    if not start_date or not end_date:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="La date de debut et la date de fin sont obligatoires",
        )

    start = datetime.combine(date.fromisoformat(start_date), time.min)
    end = datetime.combine(date.fromisoformat(end_date), time.min) + timedelta(days=1)

    if end <= start:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="La date de fin doit etre posterieure ou egale a la date de debut",
        )

    label = f"{start_date}_au_{end_date}"
    return start, end, label


def _legacy_period_range(
    period_type: Optional[str],
    year: Optional[int],
    month: Optional[int],
    week: Optional[int],
    day: Optional[str],
) -> tuple[Optional[datetime], Optional[datetime], str]:
    if not period_type:
        return None, None, "toutes_periodes"

    period = period_type.lower()
    if period == "year":
        if not year:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="L'annee est obligatoire")
        start = datetime(year, 1, 1)
        end = datetime(year + 1, 1, 1)
        return start, end, f"annee_{year}"

    if period == "month":
        if not year or not month:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="L'annee et le mois sont obligatoires",
            )
        start = datetime(year, month, 1)
        end = datetime(year + 1, 1, 1) if month == 12 else datetime(year, month + 1, 1)
        return start, end, f"mois_{year}_{month}"

    if period == "week":
        if not year or not week:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="L'annee et le numero de semaine sont obligatoires",
            )
        start_date = date.fromisocalendar(year, week, 1)
        start = datetime.combine(start_date, time.min)
        end = start + timedelta(days=7)
        return start, end, f"semaine_{year}_{week}"

    if period == "day":
        if not day:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="La date est obligatoire")
        start_date = date.fromisoformat(day)
        start = datetime.combine(start_date, time.min)
        end = start + timedelta(days=1)
        return start, end, f"jour_{day}"

    raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Periode non prise en charge")


def _alerts_for_user(db: Session, user: User) -> list[Alert]:
    stmt = (
        select(Alert)
        .options(
            joinedload(Alert.station),
            joinedload(Alert.created_by),
            joinedload(Alert.history).joinedload(AlertStatusHistory.changed_by),
            joinedload(Alert.permanent_decision).joinedload(PermanentDecision.destination_establishment),
            joinedload(Alert.permanent_decision).joinedload(PermanentDecision.permanent_user),
            joinedload(Alert.establishment_confirmation).joinedload(EstablishmentConfirmation.establishment_user),
            joinedload(Alert.attachments),
        )
        .order_by(Alert.created_at.desc())
    )

    if user.role == UserRole.AGENT:
        stmt = stmt.where(Alert.created_by_user_id == user.id)
    elif user.role == UserRole.PERMANENT:
        stmt = stmt.join(PermanentDecision, PermanentDecision.alert_id == Alert.id).where(
            PermanentDecision.permanent_user_id == user.id
        )
    elif user.role == UserRole.ETABLISSEMENT:
        stmt = stmt.outerjoin(PermanentDecision, PermanentDecision.alert_id == Alert.id).where(
            or_(
                Alert.created_by_user_id == user.id,
                PermanentDecision.destination_establishment_id == user.establishment_id,
            )
        )
    elif user.role == UserRole.SUIVI:
        stmt = stmt
    else:
        stmt = stmt.where(Alert.created_by_user_id == user.id)

    alerts = list(db.execute(stmt).unique().scalars())
    return apply_alert_collection_derived_fields(alerts)


def _online_trials_for_user(db: Session, user: User) -> list[OnlineTrialRequest]:
    stmt = (
        select(OnlineTrialRequest)
        .options(
            joinedload(OnlineTrialRequest.station),
            joinedload(OnlineTrialRequest.created_by).joinedload(User.establishment),
            joinedload(OnlineTrialRequest.departure_station),
            joinedload(OnlineTrialRequest.arrival_station),
            joinedload(OnlineTrialRequest.history).joinedload(OnlineTrialStatusHistory.changed_by),
            joinedload(OnlineTrialRequest.permanent_decision).joinedload(OnlineTrialDecision.permanent_user),
        )
        .order_by(OnlineTrialRequest.created_at.desc())
    )

    if user.role in {UserRole.AGENT, UserRole.ETABLISSEMENT, UserRole.PROJET}:
        stmt = stmt.where(OnlineTrialRequest.created_by_user_id == user.id)
    elif user.role == UserRole.PERMANENT:
        stmt = stmt.join(OnlineTrialDecision, OnlineTrialDecision.trial_id == OnlineTrialRequest.id).where(
            OnlineTrialDecision.permanent_user_id == user.id
        )

    return list(db.execute(stmt).unique().scalars())


def _filter_alerts_by_period(
    alerts: list[Alert], start: Optional[datetime], end: Optional[datetime]
) -> list[Alert]:
    if not start or not end:
        return alerts
    return [alert for alert in alerts if start <= alert.created_at.replace(tzinfo=None) < end]


def _filter_online_trials_by_period(
    trials: list[OnlineTrialRequest], start: Optional[datetime], end: Optional[datetime]
) -> list[OnlineTrialRequest]:
    if not start or not end:
        return trials
    return [trial for trial in trials if start <= trial.created_at.replace(tzinfo=None) < end]


def _filter_establishment_alerts_by_transport_scope(
    alerts: list[Alert], user: User, transport_scope: Optional[str]
) -> list[Alert]:
    if user.role != UserRole.ETABLISSEMENT:
        return alerts
    if not transport_scope:
        return alerts

    scope = transport_scope.strip().lower()
    if scope not in {"created", "reception"}:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Valeur transport_scope invalide (created ou reception)",
        )

    establishment_id = user.establishment_id
    if scope == "created":
        return [alert for alert in alerts if alert.created_by_user_id == user.id]

    if establishment_id is None:
        return []
    return [
        alert
        for alert in alerts
        if alert.permanent_decision
        and alert.permanent_decision.destination_establishment_id == establishment_id
    ]


def _format_delay_minutes(delay_minutes: Optional[int]) -> str:
    if delay_minutes is None:
        return ""
    if delay_minutes == 0:
        return "A l'heure"

    total_minutes = abs(delay_minutes)
    days, remainder = divmod(total_minutes, 1440)
    hours, minutes = divmod(remainder, 60)
    parts: list[str] = []
    if days:
        parts.append(f"{days}j")
    if hours:
        parts.append(f"{hours}h")
    if minutes or not parts:
        parts.append(f"{minutes}min")
    suffix = "retard" if delay_minutes > 0 else "avance"
    return f"{' '.join(parts)} {suffix}"


def _format_delay_minutes_tracking(delay_minutes: Optional[int]) -> str:
    if delay_minutes is None:
        return "-"
    if delay_minutes == 0:
        return "A l'heure"

    total_minutes = abs(delay_minutes)
    days = total_minutes // 1440
    hours = (total_minutes % 1440) // 60
    minutes = total_minutes % 60
    parts: list[str] = []
    if days > 0:
        parts.append(f"{days}j")
    if hours > 0:
        parts.append(f"{hours}h")
    if minutes > 0 or not parts:
        parts.append(f"{minutes}min")
    suffix = "de retard" if delay_minutes > 0 else "d'avance"
    return f"{' '.join(parts)} {suffix}"


def _split_joined_values(value: Optional[str]) -> list[str]:
    return [item.strip() for item in (value or "").split(" + ") if item.strip()]


def _parse_json_object(value: Optional[str]) -> dict[str, dict]:
    if not value:
        return {}
    try:
        parsed = json.loads(value)
    except (TypeError, ValueError):
        return {}
    if not isinstance(parsed, dict):
        return {}
    return {str(key): item for key, item in parsed.items() if isinstance(item, dict)}


def _parse_confirmed_indexes(value: Optional[str]) -> set[int]:
    indexes: set[int] = set()
    for token in (value or "").split(","):
        token = token.strip()
        if not token:
            continue
        try:
            index = int(token)
        except ValueError:
            continue
        if index >= 0:
            indexes.add(index)
    return indexes


def _format_datetime_excel(value: Optional[str]) -> str:
    if not value:
        return ""
    normalized = value.replace("Z", "+00:00")
    try:
        parsed = datetime.fromisoformat(normalized)
    except ValueError:
        return ""
    return parsed.strftime("%d/%m/%Y %H:%M")


def _format_instance_duration(minutes: int) -> str:
    if minutes <= 0:
        return "0min"
    hours, remainder = divmod(minutes, 60)
    if hours <= 0:
        return f"{remainder}min"
    if remainder == 0:
        return f"{hours}h"
    return f"{hours}h {remainder}min"


def _build_instance_observation(confirmation: dict) -> str:
    reception_status = confirmation.get("reception_status")
    instance_used_once = bool(confirmation.get("instance_used_once"))
    total_minutes = confirmation.get("en_instance_total_minutes")

    if reception_status == "EN_INSTANCE" and instance_used_once:
        started_at = (
            confirmation.get("last_instance_started_at")
            or confirmation.get("en_instance_started_at")
            or confirmation.get("confirmed_at")
        )
        started_label = _format_datetime_excel(started_at)
        return f"mis en instance le {started_label}" if started_label else ""

    if (
        isinstance(total_minutes, int)
        and total_minutes >= 0
        and instance_used_once
        and reception_status != "EN_INSTANCE"
    ):
        ended_at = (
            confirmation.get("instance_ended_at")
            or confirmation.get("reception_date")
            or confirmation.get("confirmed_at")
        )
        ended_label = _format_datetime_excel(ended_at)
        started_label = _format_datetime_excel(confirmation.get("last_instance_started_at"))
        return (
            f"etait en instance durant la periode du {started_label} au {ended_label} ({_format_instance_duration(total_minutes)})"
            if started_label and ended_label
            else ""
        )

    return ""


def _to_technicentre_code(value: Optional[str]) -> str:
    cleaned = re.sub(r"^technicentre\s+", "", (value or "").strip(), flags=re.IGNORECASE).upper()
    return cleaned if cleaned in TECHNICENTRE_CODES else ""


def _transport_ppm_state_label(value: str) -> str:
    if value == "ACCEPTEE":
        return "Acceptee"
    if value == "A_MODIFIER":
        return "A modifier"
    if value == "ANNULEE":
        return "Annulee"
    if value == "MODIFIEE":
        return "Modifiee"
    return "En attente"


def _transport_reception_label(value: str) -> str:
    if value == "VALIDEE":
        return "Validee"
    if value == "EN_INSTANCE":
        return "En instance"
    return "Non confirmee"


def _transport_status_label(value: str) -> str:
    if value == "TRAITEE_PAR_PM":
        return "Traitee par PPM"
    if value == "A_MODIFIER":
        return "A modifier"
    if value == "MODIFIEE":
        return "Modifiee"
    if value == "ANNULEE":
        return "Annulee"
    if value == "RECEPTION_PARTIELLE":
        return "Reception partielle"
    if value == "RECEPTION_COMPLETE":
        return "Reception complete"
    return "En cours de traitement"


def _online_trial_status_label(value: str) -> str:
    if value == "TRAITEE_PAR_PM":
        return "Traitee par PPM"
    if value == "A_MODIFIER":
        return "A modifier"
    if value == "MODIFIEE":
        return "Modifiee"
    if value == "ANNULEE":
        return "Annulee"
    if value == "RECEPTION_COMPLETE":
        return "Essai realise"
    if value == "RECEPTION_PARTIELLE":
        return "Traitee par PPM"
    return "En cours de traitement"


def _parse_iso_datetime_utc(value: Optional[str]) -> Optional[datetime]:
    if not value:
        return None
    normalized = value.replace("Z", "+00:00")
    try:
        parsed = datetime.fromisoformat(normalized)
    except ValueError:
        return None
    if parsed.tzinfo is None:
        return parsed
    return parsed.astimezone(timezone.utc).replace(tzinfo=None)


def _build_transport_material_workbook(
    user: User, alerts: list[Alert], period_label: str
) -> BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Suivi_Demandes"
    headers = [
        "Dossier",
        "Date demande",
        "Demandeur",
        "Site de depart",
        "Destinataire",
        "Exploitant",
        "Mode acheminement",
        "Type acheminement",
        "Etat demande (PPM)",
        "Motif PPM (modification / annulation)",
        "Statut",
        "Confirmation reception",
        "Date de reception",
        "Date systeme de reception",
        "Type materiel",
        "Serie",
        "Materiel concerne",
        "Motif",
        "Autre conditions",
        "Observation",
        "Retard",
    ]
    sheet.append(headers)

    now_utc = datetime.utcnow()
    for alert in alerts:
        types = _split_joined_values(alert.material_type)
        series = _split_joined_values(alert.material_ref)
        concerned = _split_joined_values(alert.material_concerned)
        material_count = max(len(types), len(series), len(concerned), 1)

        confirmations = _parse_json_object(
            alert.establishment_confirmation.material_confirmations
            if alert.establishment_confirmation
            else None
        )
        confirmed_indexes = _parse_confirmed_indexes(
            alert.establishment_confirmation.confirmed_material_indexes
            if alert.establishment_confirmation
            else None
        )
        ppm_decisions = _parse_json_object(
            alert.permanent_decision.material_decisions if alert.permanent_decision else None
        )
        global_ppm_status = (
            "A_MODIFIER"
            if alert.status.value == "A_MODIFIER"
            else (
                "MODIFIEE"
                if alert.status.value == "MODIFIEE"
                else ("ANNULEE" if alert.status.value == "ANNULEE" else None)
            )
        )

        destination_requested = (
            alert.requested_destination_establishment.code
            if alert.requested_destination_establishment and alert.requested_destination_establishment.code
            else (
                alert.requested_destination_establishment.name
                if alert.requested_destination_establishment
                else ""
            )
        )
        destination_retained = (
            alert.permanent_decision.destination_establishment.code
            if alert.permanent_decision
            and alert.permanent_decision.destination_establishment
            and alert.permanent_decision.destination_establishment.code
            else (
                alert.permanent_decision.destination_establishment.name
                if alert.permanent_decision and alert.permanent_decision.destination_establishment
                else ""
            )
        )
        destination_displayed = _to_technicentre_code(destination_retained or destination_requested) or "-"

        requester = (alert.created_by.full_name if alert.created_by and alert.created_by.full_name else "").strip()
        if requester.lower().startswith("technicentre "):
            requester = requester[len("technicentre ") :].strip() or requester
        requester_code = _to_technicentre_code(requester)

        request_date_value = alert.request_date or alert.created_at
        request_date_label = request_date_value.strftime("%d/%m/%Y %H:%M")

        global_ppm_reason = ""
        if alert.history:
            for item in sorted(alert.history, key=lambda it: it.changed_at, reverse=True):
                if item.status.value in {"A_MODIFIER", "ANNULEE", "MODIFIEE"} and item.note:
                    global_ppm_reason = item.note.strip()
                    break

        for index in range(material_count):
            key = str(index)
            confirmation = confirmations.get(key, {})
            is_confirmed = bool(confirmation.get("confirmed")) or index in confirmed_indexes
            raw_reception_status = confirmation.get("reception_status")
            if raw_reception_status == "VALIDEE":
                reception_state = "VALIDEE"
            elif raw_reception_status == "EN_INSTANCE":
                reception_state = "EN_INSTANCE"
            elif is_confirmed:
                reception_state = "VALIDEE"
            else:
                reception_state = "NOT_CONFIRMED"

            raw_ppm_status = ppm_decisions.get(key, {}).get("ppm_status") or global_ppm_status or "PENDING"
            is_cancelled = raw_ppm_status == "ANNULEE"
            is_in_instance = reception_state == "EN_INSTANCE"
            is_accepted_not_confirmed = raw_ppm_status == "ACCEPTEE" and reception_state == "NOT_CONFIRMED"
            is_ppm_pending = raw_ppm_status == "PENDING"
            has_final_reception = reception_state == "VALIDEE"
            row_status = (
                "ANNULEE"
                if is_cancelled
                else (
                    "EN_COURS_DE_TRAITEMENT"
                    if is_ppm_pending
                    else (
                        "RECEPTION_COMPLETE"
                        if has_final_reception
                        else ("TRAITEE_PAR_PM" if is_in_instance else ("TRAITEE_PAR_PM" if is_accepted_not_confirmed else alert.status.value))
                    )
                )
            )

            ppm_reason = ppm_decisions.get(key, {}).get("ppm_reason") or (
                global_ppm_reason if raw_ppm_status in {"A_MODIFIER", "ANNULEE", "MODIFIEE"} else ""
            )

            row_reception_date = (
                _format_datetime_excel(confirmation.get("reception_date"))
                if has_final_reception
                else ""
            )
            row_reception_system_date = ""
            if has_final_reception:
                confirmed_at_value = confirmation.get("confirmed_at")
                if isinstance(confirmed_at_value, str):
                    row_reception_system_date = _format_datetime_excel(confirmed_at_value)
                elif is_confirmed and alert.establishment_confirmation and alert.establishment_confirmation.confirmed_at:
                    row_reception_system_date = alert.establishment_confirmation.confirmed_at.strftime("%d/%m/%Y %H:%M")

            delay_value = (
                confirmation.get("delay_minutes")
                if has_final_reception and isinstance(confirmation.get("delay_minutes"), int)
                else None
            )
            if delay_value is None and is_confirmed and alert.establishment_confirmation:
                delay_value = alert.establishment_confirmation.delay_minutes

            delay_label = ""
            if isinstance(delay_value, int):
                delay_label = _format_delay_minutes_tracking(delay_value)
            elif raw_ppm_status == "ACCEPTEE" and not has_final_reception:
                ref_iso = ppm_decisions.get(key, {}).get("updated_at")
                ref_datetime = _parse_iso_datetime_utc(ref_iso) if isinstance(ref_iso, str) else None
                if not ref_datetime and alert.permanent_decision:
                    ref_datetime = alert.permanent_decision.created_at.replace(tzinfo=None)
                if ref_datetime:
                    ongoing_minutes = int((now_utc - ref_datetime).total_seconds() // 60)
                    delay_label = f"{_format_delay_minutes_tracking(ongoing_minutes)} (en cours)"

            base_observation = confirmation.get("remarks") or (
                alert.establishment_confirmation.remarks
                if is_confirmed and alert.establishment_confirmation
                else ""
            )
            instance_observation = _build_instance_observation(confirmation)
            observation = " ".join(part for part in [base_observation, instance_observation] if part).strip()

            sheet.append(
                [
                    f"#{alert.dossier_label or alert.id}",
                    request_date_label,
                    requester_code,
                    alert.station.name if alert.station else "-",
                    destination_displayed,
                    alert.maintenance_state.value,
                    alert.transport_mode,
                    alert.transport_type,
                    _transport_ppm_state_label(str(raw_ppm_status)),
                    ppm_reason,
                    _transport_status_label(row_status),
                    _transport_reception_label(reception_state),
                    row_reception_date,
                    row_reception_system_date,
                    types[index] if index < len(types) else (types[0] if types else "-"),
                    series[index] if index < len(series) else (series[0] if series else "-"),
                    concerned[index] if index < len(concerned) else (concerned[0] if concerned else "-"),
                    alert.problem_description,
                    alert.transport_conditions_initial or "-",
                    observation,
                    delay_label,
                ]
            )

    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max_length + 2, 40)

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def _build_export_workbook(user: User, alerts: list[Alert], period_label: str) -> BytesIO:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Alertes"
    summary_headers = [
        "Compte",
        "Role",
        "Periode",
        "ID alerte",
        "Date creation",
        "Gare",
        "Type materiel",
        "Reference",
        "Probleme",
        "Etat maintenance",
        "Gravite",
        "Decision agent",
        "Statut courant",
        "Conditions initiales",
        "Etablissement destinataire",
        "Reception confirmee",
        "Retard calcule",
        "Pieces jointes",
    ]
    summary_sheet.append(summary_headers)

    for alert in alerts:
        summary_sheet.append(
            [
                user.username,
                user.role.value,
                period_label,
                alert.id,
                alert.created_at.isoformat(sep=" "),
                alert.station.name if alert.station else "",
                alert.material_type,
                alert.material_ref,
                alert.problem_description,
                alert.maintenance_state.value,
                alert.severity.value,
                alert.agent_decision.value,
                _status_business_label(alert.status.value),
                alert.transport_conditions_initial,
                alert.permanent_decision.destination_establishment.name
                if alert.permanent_decision and alert.permanent_decision.destination_establishment
                else "",
                alert.establishment_confirmation.reception_date.isoformat(sep=" ")
                if alert.establishment_confirmation
                else "",
                _format_delay_minutes(
                    compute_delay_minutes(alert.establishment_confirmation, alert.permanent_decision)
                )
                if alert.establishment_confirmation
                else "",
                ", ".join(attachment.filename for attachment in alert.attachments),
            ]
        )

    if user.role == UserRole.SUIVI:
        suivi_sheet = workbook.create_sheet("Suivi_Demandes")
        suivi_headers = [
            "Dossier",
            "Date demande",
            "Demandeur",
            "Destinataire",
            "Exploitant",
            "Mode acheminement",
            "Type acheminement",
            "Etat demande (PPM)",
            "Motif PPM (modification / annulation)",
            "Statut",
            "Confirmation reception",
            "Date de reception",
            "Date systeme de reception",
            "Type materiel",
            "Serie",
            "Materiel concerne",
            "Motif",
            "Observation",
            "Retard",
        ]
        suivi_sheet.append(suivi_headers)

        now_utc = datetime.utcnow()
        for alert in alerts:
            types = _split_joined_values(alert.material_type)
            series = _split_joined_values(alert.material_ref)
            concerned = _split_joined_values(alert.material_concerned)
            material_count = max(len(types), len(series), len(concerned), 1)

            confirmations = _parse_json_object(
                alert.establishment_confirmation.material_confirmations
                if alert.establishment_confirmation
                else None
            )
            confirmed_indexes = _parse_confirmed_indexes(
                alert.establishment_confirmation.confirmed_material_indexes
                if alert.establishment_confirmation
                else None
            )
            ppm_decisions = _parse_json_object(
                alert.permanent_decision.material_decisions if alert.permanent_decision else None
            )
            global_ppm_status = "A_MODIFIER" if alert.status.value == "A_MODIFIER" else ("MODIFIEE" if alert.status.value == "MODIFIEE" else ("ANNULEE" if alert.status.value == "ANNULEE" else None))

            destination_requested = (
                alert.requested_destination_establishment.code
                if alert.requested_destination_establishment and alert.requested_destination_establishment.code
                else (alert.requested_destination_establishment.name if alert.requested_destination_establishment else "")
            )
            destination_retained = (
                alert.permanent_decision.destination_establishment.code
                if alert.permanent_decision and alert.permanent_decision.destination_establishment and alert.permanent_decision.destination_establishment.code
                else (
                    alert.permanent_decision.destination_establishment.name
                    if alert.permanent_decision and alert.permanent_decision.destination_establishment
                    else ""
                )
            )
            destination_displayed = destination_retained or destination_requested or "-"
            requester = (alert.created_by.full_name if alert.created_by and alert.created_by.full_name else "").strip()
            if requester.lower().startswith("technicentre "):
                requester = requester[len("technicentre ") :].strip() or requester
            request_date_value = alert.request_date or alert.created_at
            request_date_label = request_date_value.strftime("%d/%m/%Y %H:%M")
            global_ppm_reason = ""
            if alert.history:
                for item in sorted(alert.history, key=lambda it: it.changed_at, reverse=True):
                    if item.status.value in {"A_MODIFIER", "ANNULEE"} and item.note:
                        global_ppm_reason = item.note.strip()
                        break

            for index in range(material_count):
                key = str(index)
                confirmation = confirmations.get(key, {})
                is_confirmed = bool(confirmation.get("confirmed")) or index in confirmed_indexes
                reception_status = confirmation.get("reception_status")
                reception_label = (
                    "Validée"
                    if reception_status == "VALIDEE" or (reception_status is None and is_confirmed)
                    else ("En instance" if reception_status == "EN_INSTANCE" else "Non confirmée")
                )
                has_final_reception = reception_status == "VALIDEE" or is_confirmed
                reception_date_label = _format_datetime_excel(confirmation.get("reception_date")) if has_final_reception else ""
                reception_system_date_label = (
                    _format_datetime_excel(confirmation.get("confirmed_at"))
                    if has_final_reception
                    else ""
                )

                raw_ppm_status = (ppm_decisions.get(key, {}).get("ppm_status") or global_ppm_status or "PENDING")
                ppm_state_label = (
                    "Acceptée"
                    if raw_ppm_status == "ACCEPTEE"
                    else ("À modifier" if raw_ppm_status == "A_MODIFIER" else ("Annulée" if raw_ppm_status == "ANNULEE" else ("Modifiée" if raw_ppm_status == "MODIFIEE" else "En attente")))
                )
                ppm_reason = ppm_decisions.get(key, {}).get("ppm_reason") or (
                    global_ppm_reason if raw_ppm_status in {"A_MODIFIER", "ANNULEE", "MODIFIEE"} else ""
                )

                delay_minutes = confirmation.get("delay_minutes") if has_final_reception else None
                if not isinstance(delay_minutes, int) and is_confirmed and alert.establishment_confirmation:
                    delay_minutes = alert.establishment_confirmation.delay_minutes

                delay_label = ""
                if isinstance(delay_minutes, int):
                    delay_label = _format_delay_minutes(delay_minutes)
                elif raw_ppm_status == "ACCEPTEE" and not has_final_reception:
                    ref_iso = ppm_decisions.get(key, {}).get("updated_at")
                    ref_label = _format_datetime_excel(ref_iso) if isinstance(ref_iso, str) else ""
                    if not ref_label and alert.permanent_decision:
                        ref_label = alert.permanent_decision.created_at.strftime("%d/%m/%Y %H:%M")
                    if ref_label:
                        ref_dt = datetime.strptime(ref_label, "%d/%m/%Y %H:%M")
                        ongoing_minutes = int((now_utc - ref_dt).total_seconds() // 60)
                        delay_label = f"{_format_delay_minutes(ongoing_minutes)} (en cours)"

                base_observation = confirmation.get("remarks") or (alert.establishment_confirmation.remarks if is_confirmed and alert.establishment_confirmation else "")
                instance_observation = _build_instance_observation(confirmation)
                observation = " ".join(part for part in [base_observation, instance_observation] if part).strip()

                suivi_sheet.append(
                    [
                        f"#{alert.dossier_label or alert.id}",
                        request_date_label,
                        requester,
                        destination_displayed,
                        alert.maintenance_state.value,
                        alert.transport_mode,
                        alert.transport_type,
                        ppm_state_label,
                        ppm_reason,
                        _status_business_label(alert.status.value),
                        reception_label,
                        reception_date_label,
                        reception_system_date_label,
                        types[index] if index < len(types) else (types[0] if types else "-"),
                        series[index] if index < len(series) else (series[0] if series else "-"),
                        concerned[index] if index < len(concerned) else (concerned[0] if concerned else "-"),
                        alert.problem_description,
                        observation,
                        delay_label,
                    ]
                )

    history_sheet = workbook.create_sheet("Historique")
    history_sheet.append(["ID alerte", "Date", "Statut", "Auteur", "Note"])
    for alert in alerts:
        for item in alert.history:
            history_sheet.append(
                [
                    alert.id,
                    item.changed_at.isoformat(sep=" "),
                    _status_business_label(item.status.value),
                    item.changed_by.full_name if item.changed_by else "Systeme",
                    item.note or "",
                ]
            )

    for sheet in workbook.worksheets:
        for column_cells in sheet.columns:
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max_length + 2, 40)

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def _online_trial_parcours_label(trial: OnlineTrialRequest) -> str:
    if trial.parcours_aller and trial.parcours_retour:
        return "Aller / Retour"
    if trial.parcours_aller:
        return "Aller"
    if trial.parcours_retour:
        return "Retour"
    return "-"


def _online_trial_creator_label(trial: OnlineTrialRequest) -> str:
    if trial.created_by and trial.created_by.establishment and trial.created_by.establishment.code:
        code = _to_technicentre_code(trial.created_by.establishment.code)
        if code:
            return code
    if trial.created_by and trial.created_by.full_name:
        code = _to_technicentre_code(trial.created_by.full_name)
        if code:
            return code
    if trial.created_by and trial.created_by.username:
        code = _to_technicentre_code(trial.created_by.username)
        if code:
            return code
    return "-"


def _build_online_trials_export_workbook(
    user: User, trials: list[OnlineTrialRequest], period_label: str
) -> BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Suivi_Essais"
    headers = [
        "Dossier",
        "Date demande",
        "Createur",
        "Parcours",
        "De",
        "Vers",
        "Type materiel",
        "Serie",
        "Materiel concerne",
        "Motif",
        "Autres conditions",
        "Exploitant",
        "Statut",
        "Decision PPM",
        "Motif d'annulation/ modification",
        "Resultat",
        "Observation",
        "Retard moyen",
    ]
    sheet.append(headers)

    for trial in trials:
        material_types = _split_joined_values(trial.material_type)
        material_series = _split_joined_values(trial.material_ref)
        concerned_materials = _split_joined_values(trial.material_concerned)
        material_count = max(len(material_types), len(material_series), len(concerned_materials), 1)
        progress_by_material = _parse_json_object(trial.trial_material_progress)
        ppm_decisions = _parse_json_object(trial.permanent_decision.material_decisions if trial.permanent_decision else None)

        global_cancel_reason = ""
        if trial.history:
            for item in sorted(trial.history, key=lambda it: it.changed_at, reverse=True):
                if item.status.value in {"ANNULEE", "A_MODIFIER"} and item.note:
                    global_cancel_reason = item.note.strip()
                    break
        if not global_cancel_reason:
            global_cancel_reason = (trial.permanent_decision.comment if trial.permanent_decision else "") or ""

        per_material_delays = [
            entry.get("delay_minutes")
            for entry in progress_by_material.values()
            if isinstance(entry.get("delay_minutes"), int)
        ]
        average_delay = (
            round(sum(per_material_delays) / len(per_material_delays))
            if per_material_delays
            else None
        )

        for index in range(material_count):
            key = str(index)
            progress_entry = progress_by_material.get(key, {})
            performed = bool(progress_entry.get("performed"))
            raw_result = progress_entry.get("result")
            remarks = (
                progress_entry.get("remarks").strip()
                if isinstance(progress_entry.get("remarks"), str)
                else ""
            )
            inferred_result = (
                raw_result
                if raw_result in {"CONCLUANT", "NON_CONCLUANT"}
                else ("NON_CONCLUANT" if remarks else "CONCLUANT")
            )
            if not performed:
                result_label = "-"
                observation = "-"
            else:
                result_label = "Non Concluant" if inferred_result == "NON_CONCLUANT" else "Concluant"
                observation = remarks if inferred_result == "NON_CONCLUANT" and remarks else "-"

            ppm_reason = (
                ppm_decisions.get(key, {}).get("ppm_reason")
                if isinstance(ppm_decisions.get(key, {}).get("ppm_reason"), str)
                else ""
            )
            cancel_or_modify_reason = (
                ppm_reason
                or (
                    global_cancel_reason
                    if trial.status.value in {"ANNULEE", "A_MODIFIER", "MODIFIEE"}
                    else ""
                )
                or "-"
            )

            per_material_delay = (
                progress_entry.get("delay_minutes")
                if isinstance(progress_entry.get("delay_minutes"), int)
                else average_delay
            )

            sheet.append(
                [
                    f"#{trial.dossier_label or trial.dossier_number or trial.id}",
                    (trial.request_date or trial.created_at).strftime("%d/%m/%Y %H:%M"),
                    _online_trial_creator_label(trial),
                    _online_trial_parcours_label(trial),
                    trial.departure_station.name if trial.departure_station else (trial.station.name if trial.station else "-"),
                    trial.arrival_station.name if trial.arrival_station else (trial.station.name if trial.station else "-"),
                    material_types[index] if index < len(material_types) else (material_types[0] if material_types else "-"),
                    material_series[index] if index < len(material_series) else (material_series[0] if material_series else "-"),
                    concerned_materials[index] if index < len(concerned_materials) else "-",
                    trial.problem_description or "-",
                    trial.transport_conditions_initial or "-",
                    trial.maintenance_state.value,
                    _online_trial_status_label(trial.status.value),
                    trial.permanent_decision.decision.value if trial.permanent_decision else "-",
                    cancel_or_modify_reason,
                    result_label,
                    observation,
                    _format_delay_minutes_tracking(per_material_delay),
                ]
            )

    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max_length + 2, 40)

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


@router.get("/users", response_model=list[AdminUserRead])
def list_users(
    db: Session = Depends(get_db),
    _: User = Depends(require_roles(UserRole.ADMIN)),
) -> list[User]:
    users = list(db.execute(select(User).options(joinedload(User.establishment)).order_by(User.role, User.username)).scalars())
    return [
        user
        for user in users
        if user.role != UserRole.AGENT
        and (
            user.role != UserRole.ETABLISSEMENT
            or (user.establishment and user.establishment.code in TECHNICENTRE_CODES)
        )
    ]


@router.get("/mail-routing", response_model=AdminMailRoutingSettingsRead)
def get_mail_routing_settings(
    db: Session = Depends(get_db),
    _: User = Depends(require_roles(UserRole.ADMIN)),
) -> AdminMailRoutingSettingsRead:
    return AdminMailRoutingSettingsRead(
        permanent_pv_email=_get_setting(db, "permanent_pv_email"),
        permanent_pfl_email=_get_setting(db, "permanent_pfl_email"),
    )


@router.put("/mail-routing", response_model=AdminMailRoutingSettingsRead)
def update_mail_routing_settings(
    payload: AdminMailRoutingSettingsUpdate,
    db: Session = Depends(get_db),
    _: User = Depends(require_roles(UserRole.ADMIN)),
) -> AdminMailRoutingSettingsRead:
    payload_values = (
        payload.model_dump(exclude_unset=True)
        if hasattr(payload, "model_dump")
        else payload.dict(exclude_unset=True)
    )
    if "permanent_pv_email" in payload_values:
        _set_setting(db, "permanent_pv_email", payload.permanent_pv_email)
    if "permanent_pfl_email" in payload_values:
        _set_setting(db, "permanent_pfl_email", payload.permanent_pfl_email)
    db.commit()
    return AdminMailRoutingSettingsRead(
        permanent_pv_email=_get_setting(db, "permanent_pv_email"),
        permanent_pfl_email=_get_setting(db, "permanent_pfl_email"),
    )


@router.post("/mail-routing/test", response_model=AdminActionResponse)
def test_mail_routing_settings(
    payload: AdminMailRoutingTestPayload,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles(UserRole.ADMIN)),
) -> AdminActionResponse:
    payload_values = (
        payload.model_dump(exclude_unset=True)
        if hasattr(payload, "model_dump")
        else payload.dict(exclude_unset=True)
    )
    has_explicit_payload = bool(payload_values)

    if has_explicit_payload:
        recipients = [
            (payload_values.get("permanent_pv_email") or "").strip() or None,
            (payload_values.get("permanent_pfl_email") or "").strip() or None,
        ]
    else:
        recipients = [
            _get_setting(db, "permanent_pv_email"),
            _get_setting(db, "permanent_pfl_email"),
        ]
    if not any(recipients):
        return AdminActionResponse(
            message="Veuillez renseigner au moins une adresse email destinataire (PV/PFL)."
        )

    subject = "Test SMTP - Routage Permanent PV/PFL"
    text_body = (
        "Bonjour,\n\n"
        "Ceci est un email de test envoye depuis l'administration ONCF.\n"
        "Le routage Permanent PV/PFL est correctement configure.\n\n"
        f"Date du test: {datetime.now().strftime('%d/%m/%Y %H:%M')}\n"
        f"Declenche par: {current_user.full_name}\n"
    )
    html_body = (
        "<html><body style=\"margin:0;padding:24px;background:#f1f5f9;font-family:'Segoe UI',Arial,sans-serif;\">"
        "<div style=\"max-width:760px;margin:0 auto;background:#ffffff;border:1px solid #dbeafe;border-radius:16px;overflow:hidden;\">"
        "<div style=\"padding:18px 22px;background:linear-gradient(120deg,#0f172a,#1e3a8a);\">"
        "<p style=\"margin:0;font-size:20px;font-weight:700;color:#f8fafc;\">Test SMTP - Routage Permanent PV/PFL</p>"
        "</div>"
        "<div style=\"padding:20px 22px;\">"
        "<p style=\"margin:0 0 10px 0;font-size:14px;color:#334155;\">Bonjour,</p>"
        "<p style=\"margin:0 0 8px 0;font-size:14px;color:#334155;\">"
        "Ceci est un email de test envoye depuis l'administration ONCF.</p>"
        "<p style=\"margin:0 0 8px 0;font-size:14px;color:#334155;\">"
        "Le routage Permanent PV/PFL est correctement configure.</p>"
        f"<p style=\"margin:14px 0 0 0;font-size:12px;color:#64748b;\">Date du test: {datetime.now().strftime('%d/%m/%Y %H:%M')}</p>"
        f"<p style=\"margin:4px 0 0 0;font-size:12px;color:#64748b;\">Declenche par: {current_user.full_name}</p>"
        "</div></div></body></html>"
    )

    delivery_status, delivery_error, used_recipients = send_system_mail(
        subject=subject,
        body=text_body,
        html_body=html_body,
        sender_email=current_user.outlook_email,
        recipients=recipients,
    )
    if delivery_status != "SENT":
        return AdminActionResponse(
            message=delivery_error or "Echec envoi mail de test."
        )
    return AdminActionResponse(
        message=f"Email de test envoye avec succes ({len(used_recipients)} destinataire(s))."
    )


@router.get("/alert-form-config", response_model=AdminAlertFormConfigRead)
def get_admin_alert_form_config(
    db: Session = Depends(get_db),
    _: User = Depends(require_roles(UserRole.ADMIN)),
) -> AdminAlertFormConfigRead:
    config = get_alert_form_config(db)
    return AdminAlertFormConfigRead(
        fields={name: AdminAlertFormFieldConfig(**value) for name, value in config.items()}
    )


@router.put("/alert-form-config", response_model=AdminAlertFormConfigRead)
def update_admin_alert_form_config(
    payload: AdminAlertFormConfigUpdate,
    db: Session = Depends(get_db),
    _: User = Depends(require_roles(UserRole.ADMIN)),
) -> AdminAlertFormConfigRead:
    payload_values = payload.model_dump() if hasattr(payload, "model_dump") else payload.dict()
    config = save_alert_form_config(db, payload_values)
    db.commit()
    return AdminAlertFormConfigRead(
        fields={name: AdminAlertFormFieldConfig(**value) for name, value in config.items()}
    )


@router.get("/stations", response_model=list[StationRead])
def list_admin_stations(
    db: Session = Depends(get_db),
    _: User = Depends(require_roles(UserRole.ADMIN)),
) -> list[Station]:
    return list(db.execute(select(Station).order_by(Station.name)).scalars())


@router.post("/stations", response_model=AdminStationResponse, status_code=status.HTTP_201_CREATED)
def create_station(
    payload: AdminStationCreate,
    db: Session = Depends(get_db),
    _: User = Depends(require_roles(UserRole.ADMIN)),
) -> AdminStationResponse:
    desired_code = (payload.code or _slugify_station_code(payload.name)).strip().upper()
    existing_code = db.execute(select(Station).where(Station.code == desired_code)).scalar_one_or_none()
    if existing_code:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Code site deja utilise")

    existing_name = db.execute(select(Station).where(Station.name == payload.name.strip())).scalar_one_or_none()
    if existing_name:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Nom de site deja utilise")

    station = Station(
        code=desired_code,
        name=payload.name.strip(),
        region=payload.region.strip(),
        lat=payload.lat,
        lon=payload.lon,
    )
    db.add(station)
    db.commit()
    db.refresh(station)
    return AdminStationResponse(station=station)


@router.put("/stations/{station_id}", response_model=AdminStationResponse)
def update_station(
    station_id: int,
    payload: AdminStationUpdate,
    db: Session = Depends(get_db),
    _: User = Depends(require_roles(UserRole.ADMIN)),
) -> AdminStationResponse:
    station = db.get(Station, station_id)
    if not station:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Site introuvable")

    desired_code = (payload.code or _slugify_station_code(payload.name)).strip().upper()
    existing_code = db.execute(
        select(Station).where(Station.code == desired_code, Station.id != station_id)
    ).scalar_one_or_none()
    if existing_code:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Code site deja utilise")

    existing_name = db.execute(
        select(Station).where(Station.name == payload.name.strip(), Station.id != station_id)
    ).scalar_one_or_none()
    if existing_name:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Nom de site deja utilise")

    station.code = desired_code
    station.name = payload.name.strip()
    station.region = payload.region.strip()
    station.lat = payload.lat
    station.lon = payload.lon
    db.commit()
    db.refresh(station)
    return AdminStationResponse(station=station)


@router.delete("/stations/{station_id}", response_model=AdminActionResponse)
def delete_station(
    station_id: int,
    db: Session = Depends(get_db),
    _: User = Depends(require_roles(UserRole.ADMIN)),
) -> AdminActionResponse:
    station = db.get(Station, station_id)
    if not station:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Site introuvable")

    linked_alert = db.execute(select(Alert.id).where(Alert.station_id == station_id)).first()
    if linked_alert:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Suppression impossible: le site est deja utilise dans des demandes",
        )

    linked_revision = db.execute(select(AlertRevision.id).where(AlertRevision.station_id == station_id)).first()
    if linked_revision:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Suppression impossible: le site est present dans l'historique",
        )

    db.delete(station)
    db.commit()
    return AdminActionResponse(message="Site supprime")


@router.post("/establishments", response_model=AdminEstablishmentCreateResponse, status_code=status.HTTP_201_CREATED)
def create_establishment(
    payload: AdminEstablishmentCreate,
    db: Session = Depends(get_db),
    _: User = Depends(require_roles(UserRole.ADMIN)),
) -> AdminEstablishmentCreateResponse:
    desired_code = (payload.code or _slugify_establishment_code(payload.name)).strip().upper()
    existing_code = db.execute(select(Establishment).where(Establishment.code == desired_code)).scalar_one_or_none()
    if existing_code:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Code etablissement deja utilise")

    existing_name = db.execute(select(Establishment).where(Establishment.name == payload.name)).scalar_one_or_none()
    if existing_name:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Nom d'etablissement deja utilise")

    establishment = Establishment(
        code=desired_code,
        name=payload.name.strip(),
        city=payload.city.strip(),
        outlook_email=(payload.outlook_email or "").strip() or None,
        lat=payload.lat,
        lon=payload.lon,
    )
    db.add(establishment)
    db.commit()
    db.refresh(establishment)
    return AdminEstablishmentCreateResponse(establishment=establishment)


@router.put("/establishments/{establishment_id}", response_model=AdminEstablishmentCreateResponse)
def update_establishment(
    establishment_id: int,
    payload: AdminEstablishmentUpdate,
    db: Session = Depends(get_db),
    _: User = Depends(require_roles(UserRole.ADMIN)),
) -> AdminEstablishmentCreateResponse:
    establishment = db.get(Establishment, establishment_id)
    if not establishment:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Etablissement introuvable")

    desired_code = (payload.code or _slugify_establishment_code(payload.name)).strip().upper()
    existing_code = db.execute(
        select(Establishment).where(Establishment.code == desired_code, Establishment.id != establishment_id)
    ).scalar_one_or_none()
    if existing_code:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Code etablissement deja utilise")

    existing_name = db.execute(
        select(Establishment).where(Establishment.name == payload.name.strip(), Establishment.id != establishment_id)
    ).scalar_one_or_none()
    if existing_name:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Nom d'etablissement deja utilise")

    establishment.code = desired_code
    establishment.name = payload.name.strip()
    establishment.city = payload.city.strip()
    establishment.outlook_email = (payload.outlook_email or "").strip() or None
    establishment.lat = payload.lat
    establishment.lon = payload.lon
    db.commit()
    db.refresh(establishment)
    return AdminEstablishmentCreateResponse(establishment=establishment)


@router.post("/users", response_model=AdminUserRead, status_code=status.HTTP_201_CREATED)
def create_user(
    payload: AdminUserCreate,
    db: Session = Depends(get_db),
    _: User = Depends(require_roles(UserRole.ADMIN)),
) -> User:
    existing = db.execute(select(User).where(User.username == payload.username)).scalar_one_or_none()
    if existing:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Nom d'utilisateur deja utilise")

    _validate_establishment_assignment(db, payload.role, payload.establishment_id)
    user = User(
        username=payload.username,
        password_hash=get_password_hash(payload.password),
        role=payload.role,
        full_name=payload.full_name,
        outlook_email=(payload.outlook_email or "").strip() or None,
        establishment_id=payload.establishment_id,
    )
    db.add(user)
    db.commit()
    db.refresh(user)
    return user


@router.get("/users/{user_id}", response_model=AdminUserDetail)
def get_user_detail(
    user_id: int,
    db: Session = Depends(get_db),
    _: User = Depends(require_roles(UserRole.ADMIN)),
) -> AdminUserDetail:
    user = db.get(User, user_id)
    if not user:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Compte introuvable")
    return AdminUserDetail(user=AdminUserRead.model_validate(user), history=_user_activity_history(db, user_id))


@router.get("/users/{user_id}/export")
def export_user_excel(
    user_id: int,
    start_date: Optional[str] = Query(default=None),
    end_date: Optional[str] = Query(default=None),
    transport_scope: Optional[str] = Query(default=None),
    period_type: Optional[str] = Query(default=None),
    year: Optional[int] = Query(default=None),
    month: Optional[int] = Query(default=None),
    week: Optional[int] = Query(default=None),
    day: Optional[str] = Query(default=None),
    db: Session = Depends(get_db),
    _: User = Depends(require_roles(UserRole.ADMIN)),
) -> StreamingResponse:
    user = db.get(User, user_id)
    if not user:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Compte introuvable")

    if start_date or end_date:
        start, end, period_label = _date_range(start_date, end_date)
    else:
        start, end, period_label = _legacy_period_range(period_type, year, month, week, day)
    alerts = _filter_alerts_by_period(_alerts_for_user(db, user), start, end)
    alerts = _filter_establishment_alerts_by_transport_scope(alerts, user, transport_scope)

    if user.role == UserRole.ETABLISSEMENT:
        workbook_stream = _build_transport_material_workbook(user, alerts, period_label)
    else:
        workbook_stream = _build_export_workbook(user, alerts, period_label)

    filename = f"{user.username}_{period_label}.xlsx"
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return StreamingResponse(
        workbook_stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@router.get("/users/{user_id}/online-trials/export")
def export_user_online_trials_excel(
    user_id: int,
    start_date: Optional[str] = Query(default=None),
    end_date: Optional[str] = Query(default=None),
    period_type: Optional[str] = Query(default=None),
    year: Optional[int] = Query(default=None),
    month: Optional[int] = Query(default=None),
    week: Optional[int] = Query(default=None),
    day: Optional[str] = Query(default=None),
    db: Session = Depends(get_db),
    _: User = Depends(require_roles(UserRole.ADMIN)),
) -> StreamingResponse:
    user = db.get(User, user_id)
    if not user:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Compte introuvable")

    if start_date or end_date:
        start, end, period_label = _date_range(start_date, end_date)
    else:
        start, end, period_label = _legacy_period_range(period_type, year, month, week, day)
    trials = _filter_online_trials_by_period(_online_trials_for_user(db, user), start, end)
    workbook_stream = _build_online_trials_export_workbook(user, trials, period_label)

    filename = f"{user.username}_essais_en_ligne_{period_label}.xlsx"
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return StreamingResponse(
        workbook_stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@router.put("/users/{user_id}", response_model=AdminUserRead)
def update_user(
    user_id: int,
    payload: AdminUserUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles(UserRole.ADMIN)),
) -> User:
    user = db.get(User, user_id)
    if not user:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Compte introuvable")
    if user.id == current_user.id and payload.role != UserRole.ADMIN:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Impossible de retirer le role admin de votre propre compte",
        )

    _validate_establishment_assignment(db, payload.role, payload.establishment_id)
    user.full_name = payload.full_name
    user.role = payload.role
    user.outlook_email = (payload.outlook_email or "").strip() or None
    user.establishment_id = payload.establishment_id
    db.commit()
    db.refresh(user)
    return user


@router.put("/users/{user_id}/password", response_model=AdminActionResponse)
def update_user_password(
    user_id: int,
    payload: AdminPasswordUpdate,
    db: Session = Depends(get_db),
    _: User = Depends(require_roles(UserRole.ADMIN)),
) -> AdminActionResponse:
    user = db.get(User, user_id)
    if not user:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Compte introuvable")
    user.password_hash = get_password_hash(payload.password)
    db.commit()
    return AdminActionResponse(message="Mot de passe mis a jour")


@router.delete("/users/{user_id}", response_model=AdminActionResponse)
def delete_user(
    user_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles(UserRole.ADMIN)),
) -> AdminActionResponse:
    user = db.get(User, user_id)
    if not user:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Compte introuvable")
    if user.id == current_user.id:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Suppression de son propre compte interdite",
        )

    has_activity = any(
        [
            db.execute(select(Alert.id).where(Alert.created_by_user_id == user_id)).first(),
            db.execute(select(AlertStatusHistory.id).where(AlertStatusHistory.changed_by_user_id == user_id)).first(),
            db.execute(select(PermanentDecision.id).where(PermanentDecision.permanent_user_id == user_id)).first(),
            db.execute(
                select(EstablishmentConfirmation.id).where(
                    EstablishmentConfirmation.establishment_user_id == user_id
                )
            ).first(),
            db.execute(select(Notification.id).join(Alert).where(Alert.created_by_user_id == user_id)).first(),
            db.execute(select(OnlineTrialRequest.id).where(OnlineTrialRequest.created_by_user_id == user_id)).first(),
            db.execute(select(OnlineTrialStatusHistory.id).where(OnlineTrialStatusHistory.changed_by_user_id == user_id)).first(),
            db.execute(select(OnlineTrialDecision.id).where(OnlineTrialDecision.permanent_user_id == user_id)).first(),
        ]
    )
    if has_activity:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Compte avec historique, suppression interdite",
        )

    db.delete(user)
    db.commit()
    return AdminActionResponse(message="Compte supprime")


@router.delete("/alerts/{alert_id}", response_model=AdminActionResponse)
async def delete_alert_dossier(
    alert_id: int,
    db: Session = Depends(get_db),
    _: User = Depends(require_roles(UserRole.ADMIN)),
) -> AdminActionResponse:
    alert = db.get(Alert, alert_id)
    if not alert:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Dossier introuvable")

    dossier_root_id = alert.dossier_root_id
    dossier_number = alert.dossier_number
    alerts_to_delete = list(
        db.execute(
            select(Alert).where(
                or_(
                    Alert.id == dossier_root_id,
                    Alert.dossier_parent_id == dossier_root_id,
                    Alert.dossier_number == dossier_number,
                )
            )
        ).scalars()
    )
    if not alerts_to_delete:
        alerts_to_delete = [alert]

    alerts_to_delete.sort(key=lambda item: (item.id == dossier_root_id, -(item.dossier_iteration or 0)))
    deleted_alert_ids = [item.id for item in alerts_to_delete]

    try:
        # Delete all linked rows explicitly to avoid ORM/FK edge-cases.
        db.execute(delete(AlertStatusHistory).where(AlertStatusHistory.alert_id.in_(deleted_alert_ids)))
        db.execute(delete(PermanentDecision).where(PermanentDecision.alert_id.in_(deleted_alert_ids)))
        db.execute(delete(EstablishmentConfirmation).where(EstablishmentConfirmation.alert_id.in_(deleted_alert_ids)))
        db.execute(delete(Notification).where(Notification.alert_id.in_(deleted_alert_ids)))
        db.execute(delete(AlertAttachment).where(AlertAttachment.alert_id.in_(deleted_alert_ids)))
        db.execute(delete(AlertRevision).where(AlertRevision.alert_id.in_(deleted_alert_ids)))
        db.execute(delete(MailEvent).where(MailEvent.alert_id.in_(deleted_alert_ids)))

        # Delete alerts children first, then root.
        for item in alerts_to_delete:
            db.execute(delete(Alert).where(Alert.id == item.id))

        db.commit()
    except Exception as exc:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Echec suppression dossier: {exc}",
        ) from exc
    await manager.broadcast(
        "alerts",
        {
            "type": "alert_deleted",
            "alert_id": alert_id,
            "dossier_root_id": dossier_root_id,
            "deleted_alert_ids": deleted_alert_ids,
        },
    )

    return AdminActionResponse(
        message=f"Dossier supprime ({len(deleted_alert_ids)} demande(s) retiree(s) de la base)"
    )


@router.delete("/online-trials/{trial_id}", response_model=AdminActionResponse)
async def delete_online_trial_dossier(
    trial_id: int,
    db: Session = Depends(get_db),
    _: User = Depends(require_roles(UserRole.ADMIN)),
) -> AdminActionResponse:
    trial = db.get(OnlineTrialRequest, trial_id)
    if not trial:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Dossier d'essai introuvable")

    dossier_root_id = trial.dossier_root_id
    dossier_number = trial.dossier_number
    trials_to_delete = list(
        db.execute(
            select(OnlineTrialRequest).where(
                or_(
                    OnlineTrialRequest.id == dossier_root_id,
                    OnlineTrialRequest.dossier_parent_id == dossier_root_id,
                    OnlineTrialRequest.dossier_number == dossier_number,
                )
            )
        ).scalars()
    )
    if not trials_to_delete:
        trials_to_delete = [trial]

    trials_to_delete.sort(key=lambda item: (item.id == dossier_root_id, -(item.dossier_iteration or 0)))
    deleted_trial_ids = [item.id for item in trials_to_delete]

    try:
        db.execute(delete(OnlineTrialStatusHistory).where(OnlineTrialStatusHistory.trial_id.in_(deleted_trial_ids)))
        db.execute(delete(OnlineTrialDecision).where(OnlineTrialDecision.trial_id.in_(deleted_trial_ids)))
        db.execute(delete(OnlineTrialAttachment).where(OnlineTrialAttachment.trial_id.in_(deleted_trial_ids)))

        for item in trials_to_delete:
            db.execute(delete(OnlineTrialRequest).where(OnlineTrialRequest.id == item.id))

        db.commit()
    except Exception as exc:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Echec suppression dossier d'essai: {exc}",
        ) from exc

    await manager.broadcast(
        "alerts",
        {
            "type": "online_trial_deleted",
            "trial_id": trial_id,
            "dossier_root_id": dossier_root_id,
            "deleted_trial_ids": deleted_trial_ids,
        },
    )

    return AdminActionResponse(
        message=f"Dossier d'essai supprime ({len(deleted_trial_ids)} demande(s) retiree(s) de la base)"
    )

